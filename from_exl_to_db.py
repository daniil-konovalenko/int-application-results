import argparse

from web.app import db
from openpyxl import load_workbook

from web.app import models

parser = argparse.ArgumentParser(description='Move from excel table to sqlite3 db')
parser.add_argument('-f',
                    action='store',
                    required=True,
                    help='An xlsx file to extract data from')

if __name__ == '__main__':
    args = parser.parse_args()
    workbook = load_workbook(args.f)
    worksheet = workbook.active

    highest_column = worksheet.get_highest_column()
    highest_row = worksheet.get_highest_row()

    for row in worksheet.iter_rows(row_offset=2):
        (student_id,
         russian,
         math_unified_plan,
         math_individual_plan,
         math_final,
         final,
         ) = [cell.value for cell in row]


        results = models.Results(student_id=student_id,
                                 math_test_18=None,
                                 math=None,
                                 philology=None,
                                 history=None,
                                 science=None,
                                 biology=None,
                                 chemistry=None,
                                 english=None,
                                 geography=None,
                                 math_unified_plan=math_unified_plan,
                                 math_individual_plan=math_individual_plan,
                                 math_final=math_final,
                                 russian=russian,
                                 final=final,
                                 math_test=None
                                 )

        db.session.add(results)

    db.session.commit()
